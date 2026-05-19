"""Excel report builder with embedded thumbnails.

Layout (one row per matched pair):
  | # | Filename | Test image | Test OCR | Reference image | Reference OCR | CER | WER | Verdict |

Row colour reflects the verdict: PASS=green, WARN=yellow, FAIL=red.
A summary sheet records overall stats.

Pure openpyxl, no Excel install required to generate the file.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image

# Thumbnail target (px). Excel sizing is approximate but consistent.
THUMB_MAX_W = 260
THUMB_MAX_H = 160

VERDICT_FILL = {
    "PASS": PatternFill("solid", fgColor="C6EFCE"),
    "WARN": PatternFill("solid", fgColor="FFEB9C"),
    "FAIL": PatternFill("solid", fgColor="FFC7CE"),
}
VERDICT_FONT = {
    "PASS": Font(bold=True, color="006100"),
    "WARN": Font(bold=True, color="9C6500"),
    "FAIL": Font(bold=True, color="9C0006"),
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


@dataclass
class ReportRow:
    filename: str
    test_image: Image.Image
    ref_image: Image.Image
    test_text: str
    ref_text: str
    cer_v: float
    wer_v: float
    verdict: str


def _make_thumbnail_bytes(img: Image.Image) -> tuple[bytes, int, int]:
    """Return PNG bytes plus the resized (w, h) for Excel sizing."""
    im = img.copy()
    im.thumbnail((THUMB_MAX_W, THUMB_MAX_H), Image.LANCZOS)
    buf = io.BytesIO()
    im.save(buf, format="PNG")
    return buf.getvalue(), im.width, im.height


def _set_columns(ws):
    widths = {
        "A": 5,
        "B": 30,
        "C": 38,    # test image
        "D": 42,    # test OCR
        "E": 38,    # reference image
        "F": 42,    # reference OCR
        "G": 10,
        "H": 10,
        "I": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_workbook(
    rows: List[ReportRow],
    only_test_names: List[str],
    only_ref_names: List[str],
    title: str = "OCR Validation Report",
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = [
        "#",
        "Filename",
        "Test image",
        "Test OCR",
        "Reference image",
        "Reference OCR",
        "CER",
        "WER",
        "Verdict",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "C2"
    _set_columns(ws)

    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")

    for i, row in enumerate(rows, start=1):
        excel_row = i + 1
        ws.row_dimensions[excel_row].height = max(THUMB_MAX_H * 0.78, 110)

        ws.cell(row=excel_row, column=1, value=i).alignment = center
        ws.cell(row=excel_row, column=2, value=row.filename).alignment = wrap_top
        ws.cell(row=excel_row, column=4, value=row.test_text or "").alignment = wrap_top
        ws.cell(row=excel_row, column=6, value=row.ref_text or "").alignment = wrap_top
        ws.cell(row=excel_row, column=7, value=round(row.cer_v, 4)).alignment = center
        ws.cell(row=excel_row, column=8, value=round(row.wer_v, 4)).alignment = center
        v_cell = ws.cell(row=excel_row, column=9, value=row.verdict)
        v_cell.alignment = center
        v_cell.fill = VERDICT_FILL.get(row.verdict, PatternFill())
        v_cell.font = VERDICT_FONT.get(row.verdict, Font(bold=True))

        # apply borders + light row tint by verdict on text columns
        row_fill = VERDICT_FILL.get(row.verdict)
        for c in range(1, 10):
            cell = ws.cell(row=excel_row, column=c)
            cell.border = THIN_BORDER
            if row_fill and c in (1, 2, 4, 6, 7, 8):
                # lighter tint by reusing the same fill — Excel allows it
                cell.fill = row_fill

        for img, col_letter in ((row.test_image, "C"), (row.ref_image, "E")):
            png_bytes, w, h = _make_thumbnail_bytes(img)
            xl_img = XLImage(io.BytesIO(png_bytes))
            xl_img.width = w
            xl_img.height = h
            anchor = f"{col_letter}{excel_row}"
            ws.add_image(xl_img, anchor)

    # ---- summary sheet ----
    s = wb.create_sheet("Summary")
    n = len(rows)
    n_pass = sum(1 for r in rows if r.verdict == "PASS")
    n_warn = sum(1 for r in rows if r.verdict == "WARN")
    n_fail = sum(1 for r in rows if r.verdict == "FAIL")
    mean_cer = sum(r.cer_v for r in rows) / n if n else 0.0
    mean_wer = sum(r.wer_v for r in rows) / n if n else 0.0

    summary_rows = [
        (title, ""),
        ("", ""),
        ("Matched pairs", n),
        ("PASS  (CER <= 5%)", n_pass),
        ("WARN  (5% < CER <= 20%)", n_warn),
        ("FAIL  (CER > 20%)", n_fail),
        ("Mean CER", round(mean_cer, 4)),
        ("Mean WER", round(mean_wer, 4)),
        ("", ""),
        ("Only in test zip", len(only_test_names)),
        ("Only in reference zip", len(only_ref_names)),
    ]
    for r_idx, (k, v) in enumerate(summary_rows, start=1):
        s.cell(row=r_idx, column=1, value=k)
        s.cell(row=r_idx, column=2, value=v)
    s.cell(row=1, column=1).font = Font(bold=True, size=14)
    s.column_dimensions["A"].width = 32
    s.column_dimensions["B"].width = 18

    if only_test_names or only_ref_names:
        start = len(summary_rows) + 3
        s.cell(row=start, column=1, value="Unmatched filenames").font = Font(bold=True)
        cursor = start + 1
        if only_test_names:
            s.cell(row=cursor, column=1, value="Only in test zip:").font = Font(italic=True)
            cursor += 1
            for name in only_test_names:
                s.cell(row=cursor, column=1, value=name)
                cursor += 1
            cursor += 1
        if only_ref_names:
            s.cell(row=cursor, column=1, value="Only in reference zip:").font = Font(italic=True)
            cursor += 1
            for name in only_ref_names:
                s.cell(row=cursor, column=1, value=name)
                cursor += 1

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
