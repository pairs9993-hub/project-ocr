from __future__ import annotations

import argparse
import io
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BASE = ROOT / "artifacts" / "company_real_screens"
THUMB_MAX_W = 260
THUMB_MAX_H = 170

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def normalize_text(text: str) -> str:
    text = unicodedata.normalize("NFKC", text or "")
    return re.sub(r"\s+", " ", text).strip()


def levenshtein(left: str, right: str) -> int:
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for index, left_char in enumerate(left, 1):
        current = [index]
        for right_index, right_char in enumerate(right, 1):
            current.append(
                min(
                    previous[right_index] + 1,
                    current[-1] + 1,
                    previous[right_index - 1] + (left_char != right_char),
                )
            )
        previous = current
    return previous[-1]


def cer(reference: str, prediction: str) -> float:
    reference = normalize_text(reference)
    prediction = normalize_text(prediction)
    return levenshtein(reference, prediction) / max(1, len(reference))


def verdict(cer_value: float) -> str:
    if cer_value <= 0.05:
        return "PASS"
    if cer_value <= 0.20:
        return "WARN"
    return "FAIL"


def read_jsonl(path: Path) -> list[dict]:
    rows = []
    with path.open(encoding="utf-8") as file:
        for line_number, line in enumerate(file, 1):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"{path}:{line_number}: invalid JSONL: {exc}") from exc
    return rows


def thumbnail_bytes(path: Path) -> tuple[bytes, int, int]:
    image = Image.open(path).convert("RGB")
    image.thumbnail((THUMB_MAX_W, THUMB_MAX_H), Image.LANCZOS)
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue(), image.width, image.height


def build_rows(manifest_rows: list[dict], label_rows: list[dict]) -> tuple[list[dict], dict]:
    labels_by_name: dict[str, dict] = {}
    duplicate_labels = []
    for row in label_rows:
        name = Path(str(row.get("image", ""))).name
        if name in labels_by_name:
            duplicate_labels.append(name)
        labels_by_name[name] = row

    expected_names = [Path(str(row.get("llm_image_path", ""))).name for row in manifest_rows]
    actual_names = [Path(str(row.get("image", ""))).name for row in label_rows]
    missing = sorted(set(expected_names) - set(actual_names))
    extra = sorted(set(actual_names) - set(expected_names))

    review_rows = []
    for manifest in manifest_rows:
        image_name = Path(str(manifest["llm_image_path"])).name
        label = labels_by_name.get(image_name, {})
        llm_text = str(label.get("text", ""))
        app_text = str(manifest.get("app_ocr_text", ""))
        cer_value = cer(llm_text, app_text)
        review_rows.append(
            {
                "image": image_name,
                "image_path": str(manifest["llm_image_path"]),
                "report": manifest.get("report", ""),
                "excel_row": manifest.get("excel_row", ""),
                "source_filename": manifest.get("filename", ""),
                "app_ocr_text": app_text,
                "llm_text": llm_text,
                "llm_notes": str(label.get("notes", "")),
                "cer": cer_value,
                "verdict": verdict(cer_value),
                "needs_review": cer_value > 0.05 or bool(label.get("notes")),
            }
        )

    validation = {
        "manifest_rows": len(manifest_rows),
        "label_rows": len(label_rows),
        "missing_from_labels": missing,
        "extra_in_labels": extra,
        "duplicate_labels": sorted(set(duplicate_labels)),
    }
    return review_rows, validation


def add_review_sheet(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.active
    ws.title = "Review"
    headers = [
        "#",
        "Image name",
        "Image",
        "App OCR",
        "LLM label",
        "CER",
        "Verdict",
        "Needs review",
        "Human correction",
        "Reviewer note",
        "Report",
        "Excel row",
    ]
    ws.append(headers)
    for column_index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:L{len(rows) + 1}"

    widths = {
        "A": 5,
        "B": 44,
        "C": 38,
        "D": 42,
        "E": 42,
        "F": 10,
        "G": 12,
        "H": 14,
        "I": 42,
        "J": 30,
        "K": 34,
        "L": 10,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center")
    fills = {"PASS": PASS_FILL, "WARN": WARN_FILL, "FAIL": FAIL_FILL}

    for index, row in enumerate(rows, 1):
        excel_row = index + 1
        ws.row_dimensions[excel_row].height = 132
        values = [
            index,
            row["image"],
            "",
            row["app_ocr_text"],
            row["llm_text"],
            round(row["cer"], 4),
            row["verdict"],
            "YES" if row["needs_review"] else "",
            "",
            row["llm_notes"],
            row["report"],
            row["excel_row"],
        ]
        for column_index, value in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=column_index, value=value)
            cell.border = THIN_BORDER
            cell.alignment = center if column_index in (1, 6, 7, 8, 12) else wrap_top
            if column_index in (1, 2, 6, 7, 8) and row["verdict"] in fills:
                cell.fill = fills[row["verdict"]]

        image_path = ROOT / row["image_path"]
        png_bytes, width, height = thumbnail_bytes(image_path)
        xl_image = XLImage(io.BytesIO(png_bytes))
        xl_image.width = width
        xl_image.height = height
        ws.add_image(xl_image, f"C{excel_row}")


def add_summary_sheet(wb: Workbook, rows: list[dict], validation: dict, label_path: Path) -> None:
    ws = wb.create_sheet("Summary")
    total = len(rows)
    counts = {name: sum(1 for row in rows if row["verdict"] == name) for name in ["PASS", "WARN", "FAIL"]}
    avg_cer = sum(row["cer"] for row in rows) / total if total else 0.0
    needs_review = sum(1 for row in rows if row["needs_review"])
    summary_rows = [
        ("LLM review report", ""),
        ("Label file", str(label_path.relative_to(ROOT))),
        ("Rows", total),
        ("PASS (CER <= 5%)", counts["PASS"]),
        ("WARN (5% < CER <= 20%)", counts["WARN"]),
        ("FAIL (CER > 20%)", counts["FAIL"]),
        ("Mean app-vs-LLM CER", round(avg_cer, 4)),
        ("Needs review", needs_review),
        ("Missing labels", len(validation["missing_from_labels"])),
        ("Extra labels", len(validation["extra_in_labels"])),
        ("Duplicate labels", len(validation["duplicate_labels"])),
    ]
    for row_index, (key, value) in enumerate(summary_rows, 1):
        ws.cell(row=row_index, column=1, value=key)
        ws.cell(row=row_index, column=2, value=value)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build an Excel review report for LLM OCR labels.")
    parser.add_argument("--batch", required=True, help="Batch id, e.g. batch_001")
    parser.add_argument("--source", default="gpt", help="Label source prefix, e.g. gpt or gemini")
    parser.add_argument("--base-dir", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--manifest", type=Path, help="Override batch manifest path")
    parser.add_argument("--labels", type=Path, help="Override LLM labels JSONL path")
    parser.add_argument("--out-dir", type=Path, help="Output directory for review reports")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = args.base_dir
    manifest_path = args.manifest or base_dir / "llm_batches_20" / args.batch / "manifest.jsonl"
    label_path = args.labels or base_dir / "llm_labels" / f"{args.source}_{args.batch}.jsonl"
    out_dir = args.out_dir or base_dir / "review_reports"
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest_rows = read_jsonl(manifest_path)
    label_rows = read_jsonl(label_path)
    rows, validation = build_rows(manifest_rows, label_rows)

    workbook = Workbook()
    add_review_sheet(workbook, rows)
    add_summary_sheet(workbook, rows, validation, label_path)

    output_stem = f"{args.source}_{args.batch}_review"
    xlsx_path = out_dir / f"{output_stem}.xlsx"
    needs_review_path = out_dir / f"{output_stem}_needs_review.jsonl"
    summary_path = out_dir / f"{output_stem}_summary.json"

    workbook.save(xlsx_path)
    write_jsonl(needs_review_path, [row for row in rows if row["needs_review"]])
    summary = {
        **validation,
        "batch": args.batch,
        "source": args.source,
        "review_xlsx": str(xlsx_path.relative_to(ROOT)).replace("\\", "/"),
        "needs_review_jsonl": str(needs_review_path.relative_to(ROOT)).replace("\\", "/"),
        "rows": len(rows),
        "pass": sum(1 for row in rows if row["verdict"] == "PASS"),
        "warn": sum(1 for row in rows if row["verdict"] == "WARN"),
        "fail": sum(1 for row in rows if row["verdict"] == "FAIL"),
        "mean_cer": round(sum(row["cer"] for row in rows) / len(rows), 4) if rows else 0.0,
        "needs_review": sum(1 for row in rows if row["needs_review"]),
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()